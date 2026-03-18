#!/usr/bin/env python3
"""Excel workbook manager for per-channel cumulative reporting."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from file_utils import ensure_channel_dir, load_json, slugify

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def get_or_create_workbook(channel_name):
    channel_dir = ensure_channel_dir(channel_name)
    slug = slugify(channel_name)
    wb_path = channel_dir / f"{slug}.xlsx"

    if wb_path.exists():
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

    return wb, wb_path


def ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    return ws


def update_videos_sheet(wb, videos):
    headers = [
        "Video ID", "Title", "Published", "Duration",
        "Views", "Likes", "Transcript", "Link"
    ]
    ws = ensure_sheet(wb, "Videos", headers)

    # Build set of existing video IDs
    existing_ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            existing_ids.add(row[0])

    added = 0
    for v in videos:
        if v["video_id"] in existing_ids:
            continue
        ws.append([
            v["video_id"],
            v.get("title", ""),
            v.get("published_at", ""),
            v.get("duration", ""),
            v.get("view_count", 0),
            v.get("like_count", 0),
            "",  # transcript status filled later
            f'https://youtube.com/watch?v={v["video_id"]}',
        ])
        added += 1

    auto_width(ws)
    return added


def mark_transcripts(wb, channel_name):
    """Update transcript column based on which files exist."""
    channel_dir = ensure_channel_dir(channel_name)
    transcripts_dir = channel_dir / "transcripts"

    if "Videos" not in wb.sheetnames:
        return

    ws = wb["Videos"]
    for row in ws.iter_rows(min_row=2, max_col=8):
        vid_id = row[0].value
        if vid_id:
            transcript_path = transcripts_dir / f"{vid_id}.txt"
            row[6].value = "Yes" if transcript_path.exists() else "No"


def update_extractions_sheet(wb, channel_name, mode, topic):
    """Update extractions sheet from extraction files."""
    channel_dir = ensure_channel_dir(channel_name)
    extractions_dir = channel_dir / "extractions"

    headers = [
        "Video ID", "Title", "Mode", "Topic", "Extraction File", "Processed At"
    ]
    ws = ensure_sheet(wb, "Extractions", headers)

    # Build set of existing entries (video_id + mode + topic)
    existing = set()
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[0]:
            existing.add((row[0], row[2], row[3]))

    # Scan extraction files matching this mode/topic
    topic_slug = slugify(topic) if topic else "channel"
    pattern = f"*_{mode}_{topic_slug}.md"

    added = 0
    for f in sorted(extractions_dir.glob(pattern)):
        vid_id = f.stem.split(f"_{mode}_{topic_slug}")[0]
        key = (vid_id, mode, topic or "channel")
        if key in existing:
            continue

        # Try to find title from Videos sheet
        title = vid_id
        if "Videos" in wb.sheetnames:
            for row in wb["Videos"].iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] == vid_id:
                    title = row[1]
                    break

        ws.append([
            vid_id,
            title,
            mode,
            topic or "channel",
            f.name,
            datetime.now().isoformat(),
        ])
        added += 1

    auto_width(ws)
    return added


def update_reports_sheet(wb, mode, topic, report_file, video_count):
    headers = [
        "Timestamp", "Mode", "Topic", "Report File", "Videos Analyzed"
    ]
    ws = ensure_sheet(wb, "Reports", headers)
    ws.append([
        datetime.now().isoformat(),
        mode,
        topic or "channel",
        report_file,
        video_count,
    ])
    auto_width(ws)


def add_research_sheet(wb, mode, topic, extractions_data):
    """Add a dedicated sheet for a specific research run with detailed findings."""
    topic_slug = slugify(topic) if topic else "channel"
    sheet_name = f"{mode.capitalize()} - {(topic or 'channel')[:20]}"

    # Ensure unique sheet name
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    headers = ["Video ID", "Title", "Key Findings", "Relevance"]
    ws = ensure_sheet(wb, sheet_name, headers)

    for entry in extractions_data:
        ws.append([
            entry.get("video_id", ""),
            entry.get("title", ""),
            entry.get("summary", ""),
            entry.get("relevance", ""),
        ])

    auto_width(ws)


def main():
    parser = argparse.ArgumentParser(description="Excel workbook manager")
    parser.add_argument("--action", required=True,
                        choices=["add_videos", "mark_transcripts", "update_extractions",
                                 "add_report", "full_update"])
    parser.add_argument("--channel_name", required=True)
    parser.add_argument("--videos_file", help="Path to videos.json")
    parser.add_argument("--mode", default="research")
    parser.add_argument("--topic", default="")
    parser.add_argument("--report_file", default="")
    parser.add_argument("--video_count", type=int, default=0)
    args = parser.parse_args()

    wb, wb_path = get_or_create_workbook(args.channel_name)

    if args.action == "add_videos":
        if not args.videos_file:
            print("Error: --videos_file required", file=sys.stderr)
            sys.exit(1)
        videos = load_json(args.videos_file)
        added = update_videos_sheet(wb, videos)
        mark_transcripts(wb, args.channel_name)
        print(f"Added {added} new videos to workbook", file=sys.stderr)

    elif args.action == "mark_transcripts":
        mark_transcripts(wb, args.channel_name)
        print("Transcript status updated", file=sys.stderr)

    elif args.action == "update_extractions":
        added = update_extractions_sheet(wb, args.channel_name, args.mode, args.topic)
        print(f"Added {added} new extractions to workbook", file=sys.stderr)

    elif args.action == "add_report":
        update_reports_sheet(wb, args.mode, args.topic, args.report_file, args.video_count)
        print("Report entry added", file=sys.stderr)

    elif args.action == "full_update":
        if args.videos_file:
            videos = load_json(args.videos_file)
            update_videos_sheet(wb, videos)
        mark_transcripts(wb, args.channel_name)
        update_extractions_sheet(wb, args.channel_name, args.mode, args.topic)
        if args.report_file:
            update_reports_sheet(wb, args.mode, args.topic, args.report_file, args.video_count)
        print("Full workbook update complete", file=sys.stderr)

    wb.save(wb_path)
    print(json.dumps({"workbook": str(wb_path)}))


if __name__ == "__main__":
    main()
